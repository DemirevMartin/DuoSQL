from translations.automatic_translations import auto_tests, auto_test_names
from translations.manual_translations import manual_tests, manual_test_names
from high_level_tests import high_level_tests, high_level_test_names
import re
import openpyxl
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment, Border, Side


# Code Lines (CL) Counter
# NOTE: If a JOIN clause is present and ON is on a new line, it is NOT counted as 2nd line. It is ALWAYS counted as 1 line.
# This functionality is not implemented and has to be manually checked.
def count_code_lines(query: str):
    # This function counts the number of (non-empty) lines of code in a given query.
    query = query.strip()
    return sum(1 for line in query.splitlines() if line.strip())

# Character Counter
def count_characters(query: str):
    query = query.strip()
    # Count more than one subsequent spaces and semicolons as 1 character
    query = re.sub(r';|\s+', ' ', query)
    return len(query)

# Complexity (LoC) Counter
def count_complexity(query: str):
    """
    This function calculated the complexity as follows:
    – Base complexity: 1;
    – Joins: +1 for each JOIN clause. It is unable to count joins 
        through the FROM clause as it is much more complex to
        automate. All test queries use joins only through JOIN clauses;
    – Where: +1 for each (whole) clause;
    – View: +1 for each CREATE OR REPLACE VIEW;
    – Nested from-selects: +1 for each "FROM ( SELECT" (with flexible spaces).
    -------------------------------------------------------------------------------------------------------------------
    NOTE NOTE NOTE: This is a very basic complexity measure and does not take into account all possible SQL constructs.
    Example: if there are multiple tables in the FROM clause, it does not count them as additional complexity.
    In this project's experiments, all queries use JOINs!
    The joins through FROM and WHERE are intentionally excluded from the experiments!
    This function *does* work correctly when it comes to the queries in this project.
    -------------------------------------------------------------------------------------------------------------------
    """
    query = query.lower()
    views_used = len(re.findall(r'\bcreate or replace view\b', query))
    nested_selects = len(re.findall(r'\bfrom\s*\(\s*select\b', query))
    joins = len(re.findall(r'\bjoin\b', query))
    where_clauses = len(re.findall(r'\bwhere\b', query))
    
    complexity = 1 + joins + where_clauses + views_used + nested_selects
    return complexity

# Probabilistic Constructs Counter (only the relevant ones to this project are included)
def count_probabilistic_constructs(query: str):
    query = query.lower()
    constructs = [r'\bprob\b', r'agg_or', r'prob_Bdd', r'\&', r'_sentence', r'_dict']
    return sum(len(re.findall(construct, query)) for construct in constructs)


def run_experiment(odd_tests_name: str, even_tests_name: str):
    odd_tests_name = odd_tests_name.lower()
    even_tests_name = even_tests_name.lower()
    if odd_tests_name == even_tests_name:
        raise ValueError("Odd and even tests cannot be the same.")
    
    odd_tests, odd_test_names = (
        (manual_tests, manual_test_names) if "manual" in odd_tests_name else
        (auto_tests, auto_test_names) if "auto" in odd_tests_name else
        (high_level_tests, high_level_test_names)
    )

    even_tests, even_test_names = (
        (manual_tests, manual_test_names) if "manual" in even_tests_name else
        (auto_tests, auto_test_names) if "auto" in even_tests_name else
        (high_level_tests, high_level_test_names)
    )

    if odd_tests == even_tests:
        raise ValueError("Odd and even tests cannot be the same.")
    
    print(f"Running {odd_tests_name} tests against {even_tests_name} tests.")
    combined_tests = {}

    # Initialize combined_tests with empty lists for each key <=> Test Type
    # It doesn't matter if you take odd or even - they must both have the same keys and test number per test type
    for key in even_tests.keys():
        combined_tests[key] = []

    for key in even_tests.keys():
        for i in range(len(even_tests[key])):
            combined_tests[key].append(odd_tests[key][i])
            combined_tests[key].append(even_tests[key][i])

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Experiment Results"

    # Write header
    ws.append(["Test Type", "Test Query", "Code Lines (CL)", "Characters", "Level of Complexity (LoC)", "Probabilistic Constructs"])
    # Bold the header row
    for cell in ws[1]:
        cell.font = openpyxl.styles.Font(bold=True)

    row_idx = 2  # Start after header
    for key, tests in combined_tests.items():
        start_row = row_idx
        for test in tests:
            # NOTE: The odd tests are placed on even rows, starting from row 2, and the even tests on odd rows, starting from row 3
            # The rows increment by 2 for each test, so the first odd test is on row 2, the second odd test is on row 4, etc.; similarly for even tests
            test_name = odd_test_names[(row_idx - 2) // 2] if row_idx % 2 == 0 else even_test_names[(row_idx - 3) // 2]

            ws.append([
                key,  # Will be merged later
                test_name,
                count_code_lines(test),
                count_characters(test),
                count_complexity(test),
                count_probabilistic_constructs(test)
            ])
            row_idx += 1
        # Merge the "Test Type" cells for this group
        if row_idx - start_row > 1:
            ws.merge_cells(start_row=start_row, start_column=1, end_row=row_idx-1, end_column=1)
            merged_cell = ws.cell(row=start_row, column=1)
            merged_cell.alignment = Alignment(vertical="center", horizontal="center")
            merged_cell.font = openpyxl.styles.Font(bold=True)

    # Adjust column widths
    for col in range(1, 7):
        ws.column_dimensions[get_column_letter(col)].width = 30
        # Add borders to all cells in the column
        border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        for row in range(1, ws.max_row + 1):
            ws.cell(row=row, column=col).border = border

    filename = f"experiment_results/{odd_tests_name}_vs_{even_tests_name}.xlsx"
    wb.save(filename)
    print(f"Results saved to {filename}")



if __name__ == "__main__":
    run_experiment("manual", "high-level")
    run_experiment("manual", "auto")
